from sqlalchemy import func, or_
from openpyxl import load_workbook

from factoryos.extensions import db
from factoryos.modules.production.models import Order, QuantityReport, TimeBooking
from factoryos.models.tools import ToolMasterdata

def get_orders_overview(q, status_filter):

    query = (
        db.session.query(
            Order,
            func.coalesce(func.sum(QuantityReport.good_qty), 0).label("good_sum")
        )
        .outerjoin(QuantityReport)
        .group_by(Order.id)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Order.order_no.ilike(like),
                Order.article.ilike(like),
                Order.description.ilike(like),
                Order.tool_no.ilike(like)
            )
        )

    if status_filter:
        query = query.filter(Order.status == status_filter)

    rows = query.order_by(Order.order_no.asc()).all()

    orders = []

    for order, good_sum in rows:

        target = order.target_qty or 0
        good = good_sum or 0

        orders.append({
            "order": order,
            "good": int(good),
            "rest": max(0, target - good)
        })

    return orders, len(orders)

def create_production_order(form):

    order_no = form.get("order_no", "").strip()

    if not order_no:
        return False, "Bitte Auftragsnummer angeben."

    if Order.query.filter_by(order_no=order_no).first():
        return False, "Auftrag existiert bereits."

    try:
        target_qty = int(form.get("target_qty", 0))
    except:
        target_qty = 0

    if target_qty <= 0:
        return False, "Sollmenge muss größer 0 sein."

    o = Order(
        order_no=order_no,
        article=form.get("article"),
        article_name=form.get("article_name"),
        location=form.get("location"),
        tool_no=form.get("tool_no"),
        description=form.get("description"),
        target_qty=target_qty,
        status=form.get("status", "offen"),
        is_project=False
    )

    db.session.add(o)
    db.session.commit()

    return True, "Fertigungsauftrag angelegt."

def update_order(order_id, form):

    o = Order.query.get_or_404(order_id)

    order_no = form.get("order_no", "").strip()

    if not order_no:
        return False, "Bitte Auftragsnummer angeben."

    existing = Order.query.filter(
        Order.order_no == order_no,
        Order.id != o.id
    ).first()

    if existing:
        return False, "Auftragsnummer existiert bereits."

    try:
        target_qty = int(form.get("target_qty", 0))
    except:
        target_qty = 0

    o.order_no = order_no
    o.article = form.get("article")
    o.tool_no = form.get("tool_no")
    o.description = form.get("description")
    o.target_qty = target_qty
    o.status = form.get("status", "offen")

    db.session.commit()

    return True, "Auftrag gespeichert."

def delete_order(order_id):

    o = Order.query.get_or_404(order_id)

    running = (
        TimeBooking.query
        .filter_by(order_id=o.id)
        .filter(TimeBooking.end_time.is_(None))
        .first()
    )

    if running:
        return False, "Auftrag kann nicht gelöscht werden: aktive Buchung."

    TimeBooking.query.filter_by(order_id=o.id).delete(synchronize_session=False)
    QuantityReport.query.filter_by(order_id=o.id).delete(synchronize_session=False)

    db.session.delete(o)
    db.session.commit()

    return True, "Auftrag gelöscht."

def get_order_detail_data(order_no):

    o = Order.query.filter_by(order_no=order_no).first_or_404()

    tool = None
    if o.tool_no:
        tool = ToolMasterdata.query.filter_by(tool_no=o.tool_no).first()

    sums = (
        db.session.query(
            func.coalesce(func.sum(QuantityReport.good_qty), 0),
            func.coalesce(func.sum(QuantityReport.scrap_qty), 0)
        )
        .filter(QuantityReport.order_id == o.id)
        .first()
    )

    good = int(sums[0] or 0)
    scrap = int(sums[1] or 0)

    target = o.target_qty or 0

    return {
        "o": o,
        "tool": tool,
        "good": good,
        "scrap": scrap,
        "rest": max(0, target - good)
    }

def import_orders_from_excel(file):

    if not file:
        return False, "Bitte Excel-Datei auswählen."

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value else "" for c in ws[1]]

    required = ["order_no", "article", "description", "tool_no", "target_qty", "status"]

    for r in required:
        if r not in header:
            return False, f"Spalte fehlt: {r}"

    idx = {name: header.index(name) for name in required}

    created = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):

        order_no = str(row[idx["order_no"]] or "").strip()

        if not order_no:
            continue

        if Order.query.filter_by(order_no=order_no).first():
            skipped += 1
            continue

        try:
            target_qty = int(row[idx["target_qty"]] or 0)
        except:
            target_qty = 0

        o = Order(
            order_no=order_no,
            article=row[idx["article"]],
            description=row[idx["description"]],
            tool_no=row[idx["tool_no"]],
            target_qty=target_qty,
            status=row[idx["status"]] or "offen"
        )

        db.session.add(o)
        created += 1

    db.session.commit()

    return True, f"Import fertig: {created} neu, {skipped} übersprungen."
