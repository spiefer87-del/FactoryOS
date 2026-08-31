from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload

from factoryos.modules.masterdata.machines.models import Machine
from factoryos.modules.masterdata.machines.services.machine_import_service import EXPECTED_HEADERS
from factoryos.modules.masterdata.shared.constants import MACHINE_TYPES, MACHINE_STATUSES


def _value(value):
    return "" if value is None else value


def _format_datetime(value):
    return "" if not value else value.strftime("%d.%m.%Y %H:%M")


def _format_date(value):
    return "" if not value else value.strftime("%d.%m.%Y")


def _machine_type_label(machine_type):
    return MACHINE_TYPES.get(machine_type, machine_type or "")


def _machine_status_label(status):
    return MACHINE_STATUSES.get(status, status or "")


def _autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            50
        )


def _style_sheet(ws):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F2937"
    )
    header_font = Font(
        color="FFFFFF",
        bold=True
    )
    thin = Side(
        style="thin",
        color="E5E7EB"
    )
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws)


def _get_machines():
    return (
        Machine.query
        .options(
            joinedload(Machine.injection_molding_data),
            joinedload(Machine.created_by)
        )
        .order_by(Machine.machine_no.asc())
        .all()
    )


def _append_readable_sheet(workbook, machines):
    ws = workbook.active
    ws.title = "Maschinen"

    ws.append([
        "Maschinennummer",
        "Externe Maschinennummer",
        "Maschinenname",
        "Beschreibung",
        "Maschinentyp",
        "Maschinentyp intern",
        "Hersteller",
        "Modell",
        "Seriennummer",
        "Baujahr",
        "Steuerung",
        "Automation",
        "Betriebsstunden",
        "Letzte Wartung",
        "Nächste Wartung",
        "Standort",
        "Status",
        "Status intern",
        "Schließkraft kN",
        "Holmabstand Breite mm",
        "Holmabstand Höhe mm",
        "Min. Werkzeughöhe mm",
        "Max. Werkzeughöhe mm",
        "Öffnungshub mm",
        "Auswerferhub mm",
        "Max. Werkzeuggewicht kg",
        "Schneckendurchmesser mm",
        "Max. Schussgewicht g",
        "Max. Einspritzdruck bar",
        "Heizzonen",
        "Düsenradius mm",
        "Düsendurchmesser mm",
        "Erstellt am",
        "Erstellt von",
    ])

    for machine in machines:
        injection = machine.injection_molding_data

        ws.append([
            machine.machine_no or "",
            machine.external_machine_no or "",
            machine.name or "",
            machine.description or "",
            _machine_type_label(machine.machine_type),
            machine.machine_type or "",
            machine.manufacturer or "",
            machine.model or "",
            machine.serial_no or "",
            _value(machine.build_year),
            machine.controller_type or "",
            machine.automation_type or "",
            _value(machine.operating_hours),
            _format_datetime(machine.last_service_at),
            _format_datetime(machine.next_service_at),
            machine.location or "",
            _machine_status_label(machine.machine_status),
            machine.machine_status or "",
            _value(injection.clamping_force_kn if injection else None),
            _value(injection.tie_bar_width_mm if injection else None),
            _value(injection.tie_bar_height_mm if injection else None),
            _value(injection.min_mold_height_mm if injection else None),
            _value(injection.max_mold_height_mm if injection else None),
            _value(injection.opening_stroke_mm if injection else None),
            _value(injection.ejector_stroke_mm if injection else None),
            _value(injection.max_tool_weight_kg if injection else None),
            _value(injection.screw_diameter_mm if injection else None),
            _value(injection.max_shot_weight_g if injection else None),
            _value(injection.max_injection_pressure_bar if injection else None),
            _value(injection.heating_zones if injection else None),
            _value(injection.nozzle_radius_mm if injection else None),
            _value(injection.nozzle_diameter_mm if injection else None),
            _format_datetime(machine.created_at),
            machine.created_by.username if machine.created_by else "",
        ])

    _style_sheet(ws)


def _append_import_sheet(workbook, machines):
    ws = workbook.create_sheet("Importformat")
    ws.append(EXPECTED_HEADERS)

    for machine in machines:
        injection = machine.injection_molding_data

        values = {
            "machine_no": machine.machine_no or "",
            "external_machine_no": machine.external_machine_no or "",
            "name": machine.name or "",
            "description": machine.description or "",
            "machine_type": machine.machine_type or "",
            "manufacturer": machine.manufacturer or "",
            "model": machine.model or "",
            "serial_no": machine.serial_no or "",
            "build_year": _value(machine.build_year),
            "controller_type": machine.controller_type or "",
            "automation_type": machine.automation_type or "",
            "operating_hours": _value(machine.operating_hours),
            "last_service_at": _format_date(machine.last_service_at),
            "next_service_at": _format_date(machine.next_service_at),
            "location": machine.location or "",
            "machine_status": machine.machine_status or "",
            "clamping_force_kn": _value(injection.clamping_force_kn if injection else None),
            "tie_bar_width_mm": _value(injection.tie_bar_width_mm if injection else None),
            "tie_bar_height_mm": _value(injection.tie_bar_height_mm if injection else None),
            "min_mold_height_mm": _value(injection.min_mold_height_mm if injection else None),
            "max_mold_height_mm": _value(injection.max_mold_height_mm if injection else None),
            "opening_stroke_mm": _value(injection.opening_stroke_mm if injection else None),
            "ejector_stroke_mm": _value(injection.ejector_stroke_mm if injection else None),
            "max_tool_weight_kg": _value(injection.max_tool_weight_kg if injection else None),
            "screw_diameter_mm": _value(injection.screw_diameter_mm if injection else None),
            "max_shot_weight_g": _value(injection.max_shot_weight_g if injection else None),
            "max_injection_pressure_bar": _value(injection.max_injection_pressure_bar if injection else None),
            "heating_zones": _value(injection.heating_zones if injection else None),
            "nozzle_radius_mm": _value(injection.nozzle_radius_mm if injection else None),
            "nozzle_diameter_mm": _value(injection.nozzle_diameter_mm if injection else None),
        }

        ws.append([values.get(header, "") for header in EXPECTED_HEADERS])

    _style_sheet(ws)


def export_machines_to_excel():
    machines = _get_machines()
    workbook = Workbook()

    _append_readable_sheet(workbook, machines)
    _append_import_sheet(workbook, machines)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output
