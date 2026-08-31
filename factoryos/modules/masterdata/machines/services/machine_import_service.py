from datetime import datetime, date

from openpyxl import load_workbook

from factoryos.extensions import db

from factoryos.modules.masterdata.machines.models import (
    Machine,
    InjectionMoldingData
)

from factoryos.core.services.change_log_service import (
    log_change
)


EXPECTED_HEADERS = [
    "machine_no",
    "external_machine_no",
    "name",
    "description",
    "machine_type",
    "manufacturer",
    "model",
    "serial_no",
    "build_year",
    "controller_type",
    "automation_type",
    "operating_hours",
    "last_service_at",
    "next_service_at",
    "location",
    "machine_status",
    "clamping_force_kn",
    "tie_bar_width_mm",
    "tie_bar_height_mm",
    "min_mold_height_mm",
    "max_mold_height_mm",
    "opening_stroke_mm",
    "ejector_stroke_mm",
    "max_tool_weight_kg",
    "screw_diameter_mm",
    "max_shot_weight_g",
    "max_injection_pressure_bar",
    "heating_zones",
    "nozzle_radius_mm",
    "nozzle_diameter_mm",
]


MACHINE_TYPE_MAPPING = {
    "injection_molding": "injection_molding",
    "spritzgießmaschine": "injection_molding",
    "spritzgussmaschine": "injection_molding",
    "spritzgieß": "injection_molding",
    "spritzguss": "injection_molding",
    "milling": "milling",
    "fräsmaschine": "milling",
    "lathe": "lathe",
    "drehmaschine": "lathe",
    "assembly": "assembly",
    "montageanlage": "assembly",
    "testing": "testing",
    "prüfmaschine": "testing",
    "other": "other",
    "sonstige maschine": "other",
}


MACHINE_STATUS_MAPPING = {
    "aktiv": "aktiv",
    "active": "aktiv",
    "wartung": "wartung",
    "maintenance": "wartung",
    "defekt": "defekt",
    "defective": "defekt",
    "stillgelegt": "stillgelegt",
    "inactive": "stillgelegt",
}


def _normalize(value):
    if value is None:
        return ""
    return (
        str(value)
        .replace("\xa0", " ")
        .replace("\n", " ")
        .replace("\r", " ")
        .strip()
        .lower()
    )


def _text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value if value else None


def _integer(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    value = str(value).strip()
    if not value:
        return None
    return int(float(value.replace(",", ".")))


def _float(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    value = str(value).strip()
    if not value:
        return None
    return float(value.replace(",", "."))


def _datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    value = str(value).strip()
    for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"]:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    raise ValueError(f"Ungültiges Datum: {value}")


def _machine_type(value):
    if not value:
        return "injection_molding"
    return MACHINE_TYPE_MAPPING.get(_normalize(value))


def _machine_status(value):
    if not value:
        return "aktiv"
    return MACHINE_STATUS_MAPPING.get(_normalize(value))


def _row_dict(headers, row):
    result = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        result[header] = row[index] if index < len(row) else None
    return result


def _serialize_log_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _build_changes(instance, values, prefix=""):
    changes = {}
    for field, new_value in values.items():
        old_value = getattr(instance, field, None)
        if old_value == new_value:
            continue
        changes[f"{prefix}{field}"] = {
            "old": _serialize_log_value(old_value),
            "new": _serialize_log_value(new_value)
        }
    return changes


def import_machines_from_excel(file, user_id):
    workbook = load_workbook(file, data_only=True)
    worksheet = workbook.active

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_normalize(header) if header else "" for header in header_row]

    if "machine_no" not in headers:
        raise ValueError("Die Spalte 'machine_no' fehlt.")

    created = 0
    updated = 0
    errors = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and str(value).strip() != "" for value in row):
            continue

        data = _row_dict(headers, row)
        machine_no = _text(data.get("machine_no"))

        if not machine_no:
            errors.append({"row": row_index, "machine_no": "-", "reason": "Maschinennummer fehlt"})
            continue

        raw_type = data.get("machine_type")
        machine_type = _machine_type(raw_type)
        if not machine_type:
            errors.append({
                "row": row_index,
                "machine_no": machine_no,
                "reason": f"Unbekannter Maschinentyp: '{raw_type}'"
            })
            continue

        raw_status = data.get("machine_status")
        machine_status = _machine_status(raw_status)
        if not machine_status:
            errors.append({
                "row": row_index,
                "machine_no": machine_no,
                "reason": f"Unbekannter Maschinenstatus: '{raw_status}'"
            })
            continue

        try:
            with db.session.begin_nested():
                machine = Machine.query.filter_by(machine_no=machine_no).first()
                is_new = machine is None

                if is_new:
                    machine = Machine(machine_no=machine_no, created_by_id=user_id)
                    db.session.add(machine)
                    db.session.flush()

                machine_values = {
                    "external_machine_no": _text(data.get("external_machine_no")),
                    "name": _text(data.get("name")),
                    "description": _text(data.get("description")),
                    "machine_type": machine_type,
                    "manufacturer": _text(data.get("manufacturer")),
                    "model": _text(data.get("model")),
                    "serial_no": _text(data.get("serial_no")),
                    "build_year": _integer(data.get("build_year")),
                    "controller_type": _text(data.get("controller_type")),
                    "automation_type": _text(data.get("automation_type")),
                    "operating_hours": _integer(data.get("operating_hours")),
                    "last_service_at": _datetime(data.get("last_service_at")),
                    "next_service_at": _datetime(data.get("next_service_at")),
                    "location": _text(data.get("location")),
                    "machine_status": machine_status,
                }

                changes = _build_changes(machine, machine_values)
                for field, value in machine_values.items():
                    setattr(machine, field, value)

                if machine_type == "injection_molding":
                    injection_values = {
                        "clamping_force_kn": _float(data.get("clamping_force_kn")),
                        "tie_bar_width_mm": _float(data.get("tie_bar_width_mm")),
                        "tie_bar_height_mm": _float(data.get("tie_bar_height_mm")),
                        "min_mold_height_mm": _float(data.get("min_mold_height_mm")),
                        "max_mold_height_mm": _float(data.get("max_mold_height_mm")),
                        "opening_stroke_mm": _float(data.get("opening_stroke_mm")),
                        "ejector_stroke_mm": _float(data.get("ejector_stroke_mm")),
                        "max_tool_weight_kg": _float(data.get("max_tool_weight_kg")),
                        "screw_diameter_mm": _float(data.get("screw_diameter_mm")),
                        "max_shot_weight_g": _float(data.get("max_shot_weight_g")),
                        "max_injection_pressure_bar": _float(data.get("max_injection_pressure_bar")),
                        "heating_zones": _integer(data.get("heating_zones")),
                        "nozzle_radius_mm": _float(data.get("nozzle_radius_mm")),
                        "nozzle_diameter_mm": _float(data.get("nozzle_diameter_mm")),
                    }

                    injection = machine.injection_molding_data
                    if injection is None:
                        injection = InjectionMoldingData(machine=machine)
                        db.session.add(injection)

                    changes.update(_build_changes(injection, injection_values, prefix="injection."))
                    for field, value in injection_values.items():
                        setattr(injection, field, value)

                elif machine.injection_molding_data:
                    changes["injection_molding_data"] = {"old": "vorhanden", "new": None}
                    db.session.delete(machine.injection_molding_data)

                db.session.flush()

                if is_new:
                    log_change(
                        entity_type="machine",
                        entity_id=machine.id,
                        entity_name=machine.machine_no,
                        action="import",
                        changes={"machine_no": {"old": None, "new": machine.machine_no}},
                        category="masterdata"
                    )
                    created += 1
                else:
                    if changes:
                        log_change(
                            entity_type="machine",
                            entity_id=machine.id,
                            entity_name=machine.machine_no,
                            action="import",
                            changes=changes,
                            category="masterdata"
                        )
                    updated += 1

        except Exception as error:
            errors.append({"row": row_index, "machine_no": machine_no, "reason": str(error)})

    db.session.commit()
    return {"created": created, "updated": updated, "errors": errors}
