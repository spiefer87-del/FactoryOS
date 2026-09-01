from datetime import date, datetime

from openpyxl import load_workbook

from factoryos.extensions import db
from factoryos.modules.masterdata.tools.models import Tool


EXPECTED_HEADERS = [
    "tool_no",
    "external_tool_no",
    "name",
    "description",
    "built_by",
    "build_year",
    "shot_counter",
    "last_service_at",
    "location",
    "tool_status",
    "cavities",
    "core_pulls",
    "hotrunner_zones",
    "tool_weight_kg",
    "tool_length_mm",
    "tool_width_mm",
    "tool_height_mm",
    "centering_nozzle_side",
    "centering_ejector_side",
    "ejector_connection",
    "demolding_type",
    "automation_type",
    "has_conversion_kit",
]


HEADER_ALIASES = {
    "werkzeugnummer": "tool_no",
    "externe werkzeugnummer": "external_tool_no",
    "werkzeugname": "name",
    "beschreibung": "description",
    "hersteller": "built_by",
    "baujahr": "build_year",
    "schusszähler": "shot_counter",
    "schusszaehler": "shot_counter",
    "letzte wartung": "last_service_at",
    "standort": "location",
    "status": "tool_status",
    "status intern": "tool_status",
    "kavitäten": "cavities",
    "kavitaeten": "cavities",
    "kernzüge": "core_pulls",
    "kernzuege": "core_pulls",
    "heißkanalzonen": "hotrunner_zones",
    "heisskanalzonen": "hotrunner_zones",
    "gewicht kg": "tool_weight_kg",
    "länge mm": "tool_length_mm",
    "laenge mm": "tool_length_mm",
    "breite mm": "tool_width_mm",
    "höhe mm": "tool_height_mm",
    "hoehe mm": "tool_height_mm",
    "zentrierung düsenseite": "centering_nozzle_side",
    "zentrierung duesenseite": "centering_nozzle_side",
    "zentrierung auswerferseite": "centering_ejector_side",
    "auswerferanschluss": "ejector_connection",
    "entformung": "demolding_type",
    "automation": "automation_type",
    "umbausatz": "has_conversion_kit",
}


STATUS_MAPPING = {
    "aktiv": "aktiv",
    "active": "aktiv",
    "wartung": "wartung",
    "maintenance": "wartung",
    "defekt": "defekt",
    "defective": "defekt",
    "beim kunden": "external",
    "external": "external",
    "verschrottet": "scrapped",
    "scrapped": "scrapped",
}


FIELD_CONVERTERS = {
    "external_tool_no": "text",
    "name": "text",
    "description": "text",
    "built_by": "text",
    "build_year": "integer",
    "shot_counter": "integer",
    "last_service_at": "datetime",
    "location": "text",
    "cavities": "integer",
    "core_pulls": "integer",
    "hotrunner_zones": "integer",
    "tool_weight_kg": "float",
    "tool_length_mm": "float",
    "tool_width_mm": "float",
    "tool_height_mm": "float",
    "centering_nozzle_side": "text",
    "centering_ejector_side": "text",
    "ejector_connection": "text",
    "demolding_type": "text",
    "automation_type": "text",
    "has_conversion_kit": "boolean",
}


def _normalize(value):
    if value is None:
        return ""

    return (
        str(value)
        .replace("\xa0", " ")
        .replace("\n", " ")
        .replace("\r", " ")
        .strip()
        .lower()
    )


def _canonical_header(value):
    normalized = _normalize(value)

    if normalized in EXPECTED_HEADERS:
        return normalized

    return HEADER_ALIASES.get(normalized, "")


def _text(value):
    if value is None:
        return None

    value = str(value).strip()
    return value if value else None


def _integer(value):
    if value in (None, ""):
        return None

    if isinstance(value, (int, float)):
        return int(value)

    value = str(value).strip()

    if not value:
        return None

    return int(float(value.replace(",", ".")))


def _float(value):
    if value in (None, ""):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip()

    if not value:
        return None

    return float(value.replace(",", "."))


def _datetime(value):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    value = str(value).strip()

    try:
        return datetime.fromisoformat(value)
    except ValueError:
        pass

    for date_format in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(value, date_format)
        except ValueError:
            pass

    raise ValueError(f"Ungültiges Datum: {value}")


def _boolean(value):
    if value in (None, ""):
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)

    normalized = _normalize(value)

    if normalized in {"ja", "yes", "true", "wahr", "1", "x"}:
        return True

    if normalized in {"nein", "no", "false", "falsch", "0"}:
        return False

    raise ValueError(f"Ungültiger Ja/Nein-Wert: {value}")


CONVERTERS = {
    "text": _text,
    "integer": _integer,
    "float": _float,
    "datetime": _datetime,
    "boolean": _boolean,
}


def _tool_status(value):
    if value in (None, ""):
        return None

    return STATUS_MAPPING.get(_normalize(value))


def _row_dict(headers, row):
    result = {}

    for index, header in enumerate(headers):
        if not header:
            continue

        value = row[index] if index < len(row) else None

        # Im alten Export gibt es "Status" und "Status intern". Der spätere,
        # nicht leere Wert ist die verlässlichere Importquelle.
        if header in result and value in (None, ""):
            continue

        result[header] = value

    return result


def _find_import_worksheet(workbook):
    worksheets = [workbook.active]
    worksheets.extend(
        worksheet
        for worksheet in workbook.worksheets
        if worksheet is not workbook.active
    )

    for worksheet in worksheets:
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        headers = [_canonical_header(header) for header in header_row]

        if "tool_no" in headers:
            return worksheet, headers

    raise ValueError(
        "Die Spalte 'tool_no' bzw. 'Werkzeugnummer' fehlt. "
        "Bitte die aktuelle Importvorlage oder einen Werkzeug-Export verwenden."
    )


def import_tools_from_excel(file):
    workbook = load_workbook(file, data_only=True)
    worksheet, headers = _find_import_worksheet(workbook)
    available_headers = set(headers)

    created = 0
    updated = 0
    errors = []

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(
            value is not None and str(value).strip() != ""
            for value in row
        ):
            continue

        data = _row_dict(headers, row)
        tool_no = _text(data.get("tool_no"))

        if not tool_no:
            errors.append({
                "row": row_index,
                "tool_no": "-",
                "reason": "Werkzeugnummer fehlt",
            })
            continue

        mapped_status = None

        if "tool_status" in available_headers:
            raw_status = data.get("tool_status")
            mapped_status = _tool_status(raw_status)

            if raw_status not in (None, "") and not mapped_status:
                errors.append({
                    "row": row_index,
                    "tool_no": tool_no,
                    "reason": f"Unbekannter Status: '{raw_status}'",
                })
                continue

        try:
            with db.session.begin_nested():
                tool = Tool.query.filter_by(tool_no=tool_no).first()
                is_new = tool is None

                if is_new:
                    tool = Tool(
                        tool_no=tool_no,
                        tool_status=mapped_status or "aktiv",
                    )
                    db.session.add(tool)

                for field, converter_name in FIELD_CONVERTERS.items():
                    if field not in available_headers:
                        continue

                    converter = CONVERTERS[converter_name]
                    setattr(tool, field, converter(data.get(field)))

                if mapped_status:
                    tool.tool_status = mapped_status

                db.session.flush()

            if is_new:
                created += 1
            else:
                updated += 1

        except Exception as error:
            errors.append({
                "row": row_index,
                "tool_no": tool_no,
                "reason": str(error),
            })

    db.session.commit()

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
    }
