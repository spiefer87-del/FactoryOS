from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload

from factoryos.modules.masterdata.shared.constants import TOOL_STATUSES
from factoryos.modules.masterdata.tools.models import Tool
from factoryos.modules.masterdata.tools.services.tool_import_service import (
    EXPECTED_HEADERS,
)


def _value(value):
    return "" if value is None else value


def _format_datetime(value):
    if not value:
        return ""

    return value.strftime("%d.%m.%Y %H:%M")


def _bool_label(value):
    return "Ja" if value else "Nein"


def _tool_status_label(status):
    return TOOL_STATUSES.get(status, status or "")


def _join_article_numbers(tool):
    if not tool.articles:
        return ""

    return ", ".join(
        article.article_no
        for article in tool.articles
        if article.article_no
    )


def _join_article_names(tool):
    if not tool.articles:
        return ""

    return ", ".join(
        article.article_name
        for article in tool.articles
        if article.article_name
    )


def _dimensions(tool):
    if (
        tool.tool_length_mm is None
        and tool.tool_width_mm is None
        and tool.tool_height_mm is None
    ):
        return ""

    return (
        f"{_value(tool.tool_length_mm) or '-'} × "
        f"{_value(tool.tool_width_mm) or '-'} × "
        f"{_value(tool.tool_height_mm) or '-'} mm"
    )


def _autosize_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            50,
        )


def _style_sheet(worksheet):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F2937",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )
    thin = Side(
        style="thin",
        color="E5E7EB",
    )
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)


def _get_tools():
    return (
        Tool.query
        .options(
            joinedload(Tool.articles),
            joinedload(Tool.created_by),
            joinedload(Tool.images),
        )
        .order_by(Tool.tool_no.asc())
        .all()
    )


def _append_import_sheet(workbook, tools):
    worksheet = workbook.active
    worksheet.title = "Werkzeuge"
    worksheet.append(EXPECTED_HEADERS)

    for tool in tools:
        values = {
            "tool_no": tool.tool_no or "",
            "external_tool_no": tool.external_tool_no or "",
            "name": tool.name or "",
            "description": tool.description or "",
            "built_by": tool.built_by or "",
            "build_year": _value(tool.build_year),
            "shot_counter": _value(tool.shot_counter),
            "last_service_at": _value(tool.last_service_at),
            "location": tool.location or "",
            "tool_status": tool.tool_status or "aktiv",
            "cavities": _value(tool.cavities),
            "core_pulls": _value(tool.core_pulls),
            "hotrunner_zones": _value(tool.hotrunner_zones),
            "tool_weight_kg": _value(tool.tool_weight_kg),
            "tool_length_mm": _value(tool.tool_length_mm),
            "tool_width_mm": _value(tool.tool_width_mm),
            "tool_height_mm": _value(tool.tool_height_mm),
            "centering_nozzle_side": tool.centering_nozzle_side or "",
            "centering_ejector_side": tool.centering_ejector_side or "",
            "ejector_connection": tool.ejector_connection or "",
            "demolding_type": tool.demolding_type or "",
            "automation_type": tool.automation_type or "",
            "has_conversion_kit": _bool_label(tool.has_conversion_kit),
        }

        worksheet.append([
            values.get(header, "")
            for header in EXPECTED_HEADERS
        ])

    last_service_column = EXPECTED_HEADERS.index("last_service_at") + 1

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(
            row=row_index,
            column=last_service_column,
        ).number_format = "DD.MM.YYYY HH:MM"

    _style_sheet(worksheet)


def _append_readable_sheet(workbook, tools):
    worksheet = workbook.create_sheet("Übersicht")

    worksheet.append([
        "Werkzeugnummer",
        "Externe Werkzeugnummer",
        "Werkzeugname",
        "Beschreibung",
        "Status",
        "Status intern",
        "Standort",
        "Artikelnummern",
        "Artikelnamen",
        "Hersteller",
        "Baujahr",
        "Schusszähler",
        "Letzte Wartung",
        "Kavitäten",
        "Kernzüge",
        "Heißkanalzonen",
        "Gewicht kg",
        "Länge mm",
        "Breite mm",
        "Höhe mm",
        "Abmessungen",
        "Zentrierung Düsenseite",
        "Zentrierung Auswerferseite",
        "Auswerferanschluss",
        "Entformung",
        "Automation",
        "Umbausatz",
        "Bilder Anzahl",
        "Erstellt am",
        "Erstellt von",
    ])

    for tool in tools:
        worksheet.append([
            tool.tool_no or "",
            tool.external_tool_no or "",
            tool.name or "",
            tool.description or "",
            _tool_status_label(tool.tool_status),
            tool.tool_status or "",
            tool.location or "",
            _join_article_numbers(tool),
            _join_article_names(tool),
            tool.built_by or "",
            _value(tool.build_year),
            _value(tool.shot_counter),
            _format_datetime(tool.last_service_at),
            _value(tool.cavities),
            _value(tool.core_pulls),
            _value(tool.hotrunner_zones),
            _value(tool.tool_weight_kg),
            _value(tool.tool_length_mm),
            _value(tool.tool_width_mm),
            _value(tool.tool_height_mm),
            _dimensions(tool),
            tool.centering_nozzle_side or "",
            tool.centering_ejector_side or "",
            tool.ejector_connection or "",
            tool.demolding_type or "",
            tool.automation_type or "",
            _bool_label(tool.has_conversion_kit),
            len(tool.images) if tool.images else 0,
            _format_datetime(tool.created_at),
            tool.created_by.username if tool.created_by else "",
        ])

    _style_sheet(worksheet)


def _append_article_sheet(workbook, tools):
    worksheet = workbook.create_sheet("Artikelzuordnung")

    worksheet.append([
        "Werkzeugnummer",
        "Externe Werkzeugnummer",
        "Werkzeugname",
        "Artikelnummer",
        "Artikelname",
    ])

    for tool in tools:
        if not tool.articles:
            worksheet.append([
                tool.tool_no or "",
                tool.external_tool_no or "",
                tool.name or "",
                "",
                "",
            ])
            continue

        for article in tool.articles:
            worksheet.append([
                tool.tool_no or "",
                tool.external_tool_no or "",
                tool.name or "",
                article.article_no or "",
                article.article_name or "",
            ])

    _style_sheet(worksheet)


def export_tools_to_excel():
    tools = _get_tools()
    workbook = Workbook()

    # Das erste Blatt entspricht exakt dem Importformat. Dadurch kann die
    # komplette Exportdatei ohne Umbau wieder hochgeladen werden.
    _append_import_sheet(workbook, tools)
    _append_readable_sheet(workbook, tools)
    _append_article_sheet(workbook, tools)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output
