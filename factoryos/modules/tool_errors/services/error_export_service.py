from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import or_
from sqlalchemy.orm import joinedload

from factoryos.modules.tool_errors.models import ToolError
from factoryos.modules.tool_errors.constants import WORKFLOW_STATUSES
from factoryos.modules.masterdata.shared.constants import TOOL_STATUSES


def _format_datetime(value):

    if not value:
        return ""

    return value.strftime("%d.%m.%Y %H:%M")


def _bool_label(value):

    return "Ja" if value else "Nein"


def _workflow_label(status):

    return WORKFLOW_STATUSES.get(status, status or "")


def _tool_status_label(status):

    return TOOL_STATUSES.get(status, status or "")


def _autosize_columns(ws):

    for column_cells in ws.columns:

        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            if cell.value is None:
                continue

            max_length = max(
                max_length,
                len(str(cell.value))
            )

        ws.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            45
        )


def _style_sheet(ws):

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F2937"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    thin = Side(
        style="thin",
        color="E5E7EB"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.border = border

    for row in ws.iter_rows(min_row=2):

        for cell in row:

            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    _autosize_columns(ws)


def _get_tool_errors(include_history=False):

    query = (
        ToolError.query
        .options(
            joinedload(ToolError.tool),
            joinedload(ToolError.reported_by),
            joinedload(ToolError.released_by)
        )
    )

    if not include_history:

        query = query.filter(
            or_(
                ToolError.is_current.is_(True),
                ToolError.is_current.is_(None)
            )
        )

    return (
        query
        .order_by(
            ToolError.error_no.asc(),
            ToolError.revision.asc(),
            ToolError.id.asc()
        )
        .all()
    )


def export_tool_errors_to_excel(include_history=False):

    errors = _get_tool_errors(
        include_history=include_history
    )

    wb = Workbook()

    # =====================================================
    # SHEET 1: VOLLSTÄNDIGER EXPORT
    # =====================================================

    ws = wb.active
    ws.title = "Tool Errors"

    ws.append([
        "FM-Nr.",
        "Revision",
        "Aktuelle Revision",
        "Workflow",
        "Workflow intern",

        "Werkzeugnummer",
        "Externe Werkzeugnummer",
        "Werkzeugname",
        "Werkzeugstatus",

        "Fehlerart",
        "Beschreibung",

        "Auftrag",
        "Maschine",

        "Erstellt am",
        "Erstellt von",

        "Freigegeben am",
        "Freigegeben von",
    ])

    for error in errors:

        tool = error.tool

        ws.append([
            error.error_no or "",
            error.revision or 1,
            _bool_label(error.is_current),
            _workflow_label(error.workflow_status),
            error.workflow_status or "",

            tool.tool_no if tool else "",
            tool.external_tool_no if tool else "",
            tool.name if tool else "",
            _tool_status_label(tool.tool_status) if tool else "",

            error.error_type or "",
            error.description or "",

            error.order_id or "",
            error.machine_id or "",

            _format_datetime(error.created_at),
            error.reported_by.username if error.reported_by else "",

            _format_datetime(error.released_at),
            error.released_by.username if error.released_by else "",
        ])

    _style_sheet(ws)


    # =====================================================
    # SHEET 2: IMPORT-FORMAT
    # =====================================================

    ws_import = wb.create_sheet("Import Format")

    ws_import.append([
        "error_no",
        "tool_no",
        "error_type",
        "description",
        "tool_status",
        "order_id",
        "machine_id"
    ])

    for error in errors:

        tool = error.tool

        ws_import.append([
            error.error_no or "",
            tool.tool_no if tool else "",
            error.error_type or "",
            error.description or "",
            tool.tool_status if tool else "",
            error.order_id or "",
            error.machine_id or "",
        ])

    _style_sheet(ws_import)


    output = BytesIO()

    wb.save(output)
    output.seek(0)

    return output
