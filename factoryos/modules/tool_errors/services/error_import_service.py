from datetime import datetime

from openpyxl import load_workbook

from factoryos.extensions import db
from factoryos.core.services.change_log_service import log_change

from factoryos.modules.masterdata.tools.models import Tool
from factoryos.modules.tool_errors.models import ToolError
from factoryos.modules.tool_errors.services.tool_error_storage_service import (
    create_tool_error_folders,
)


STATUS_MAPPING = {
    "aktiv": "aktiv",
    "wartung": "wartung",
    "defekt": "defekt",

    "beim kunden": "external",
    "external": "external",

    "verschrottet": "scrapped",
    "scrapped": "scrapped"
}


def normalize_status(value):

    if not value:
        return None

    return (
        str(value)
        .replace("\xa0", " ")
        .replace("\n", "")
        .replace("\r", "")
        .strip()
        .lower()
    )


def import_errors_from_excel(file, user_id):

    wb = load_workbook(file)

    ws = wb.active

    created = 0
    errors = []

    for row_index, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2
    ):

        if not row:
            continue

        # ==========================================
        # FM NUMMER
        # ==========================================

        error_no = str(row[0]).strip() if row[0] else ""

        if not error_no:

            errors.append({
                "row": row_index,
                "reason": "Fehlernummer fehlt"
            })

            continue

        # ==========================================
        # WERKZEUGNUMMER
        # ==========================================

        tool_no = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if not tool_no:

            errors.append({
                "row": row_index,
                "error_no": error_no,
                "reason": "Werkzeugnummer fehlt"
            })

            continue

        tool = Tool.query.filter_by(
            tool_no=tool_no
        ).first()

        if not tool:

            errors.append({
                "row": row_index,
                "error_no": error_no,
                "tool_no": tool_no,
                "reason": "Werkzeug nicht gefunden"
            })

            continue

        # ==========================================
        # FEHLERART / BESCHREIBUNG / STATUS
        # ==========================================

        error_type = (
            str(row[2]).strip()
            if len(row) > 2 and row[2]
            else ""
        )

        description = (
            str(row[3]).strip()
            if len(row) > 3 and row[3]
            else ""
        )

        tool_status_raw = (
            str(row[4]).strip()
            if len(row) > 4 and row[4]
            else None
        )

        # ==========================================
        # DOPPELTE FM NUMMER
        # ==========================================

        existing = ToolError.query.filter_by(
            error_no=error_no
        ).first()

        if existing:

            errors.append({
                "row": row_index,
                "error_no": error_no,
                "tool_no": tool_no,
                "reason": "FM-Nummer existiert bereits"
            })

            continue

        # ==========================================
        # STATUS MAPPING
        # ==========================================

        mapped_status = None

        if tool_status_raw:

            normalized_status = normalize_status(tool_status_raw)

            mapped_status = STATUS_MAPPING.get(normalized_status)

            if not mapped_status:

                errors.append({
                    "row": row_index,
                    "error_no": error_no,
                    "tool_no": tool_no,
                    "reason": f"Ungültiger Werkzeugstatus: {tool_status_raw}"
                })

                continue

        # ==========================================
        # TOOL ERROR ANLEGEN
        # Wichtig:
        # workflow_status, revision und is_current
        # explizit setzen, damit Workflow-Buttons funktionieren.
        # ==========================================

        error = ToolError(
            error_no=error_no,

            tool_id=tool.id,

            error_type=error_type,
            description=description,

            reported_by_id=user_id,
            created_at=datetime.utcnow(),

            workflow_status="draft",
            revision=1,
            parent_error_id=None,
            is_current=True
        )

        db.session.add(error)
        db.session.flush()
        create_tool_error_folders(error)

        # ==========================================
        # TOOL STATUS AKTUALISIEREN
        # ==========================================

        changes = {
            "Werkzeug": {
                "old": None,
                "new": tool.tool_no
            },
            "error_type": {
                "old": None,
                "new": error_type or "-"
            },
            "description": {
                "old": None,
                "new": description or "-"
            },
            "workflow_status": {
                "old": None,
                "new": "Entwurf"
            },
            "revision": {
                "old": None,
                "new": 1
            }
        }

        if mapped_status:

            old_status = tool.tool_status

            if old_status != mapped_status:

                tool.tool_status = mapped_status

                changes["Werkzeug Status"] = {
                    "old": old_status,
                    "new": mapped_status
                }

        log_change(
            entity_type="tool_error",
            entity_id=error.id,
            entity_name=f"{error.error_no} Rev. {error.revision} ({tool.tool_no})",
            action="import",
            changes=changes,
            category="production"
        )

        created += 1

    db.session.commit()

    return {
        "created": created,
        "errors": errors
    }
